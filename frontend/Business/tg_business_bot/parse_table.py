import openpyxl

from tg_business_bot.Models.Menu import Menu
from tg_business_bot.Models.Category import Category
from tg_business_bot.Models.Dish import Dish
from ApiClient import ApiClient


def parse_menu_from_excel(file_path: str, menu: Menu) -> Menu:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    is_drinks = "Объем (мл)" in headers
    is_dishes = "Вес (гр)" in headers

    category_api_client = ApiClient(Category)

    categories = {}
    menu_items = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_has_info = any(row[1:12])
        if not row_has_info:
            continue

        category_name = row[1]
        name = row[2]
        description = row[3]
        price = row[4]
        weight_or_volume = row[5]
        cooking_time_minutes = row[10]

        if category_name not in categories:
            category = Category(menu_id=menu.restaurant_id, name=category_name)
            category.id = category_api_client.post(category)
            categories[category_name] = category

        category = categories[category_name]

        if is_drinks:
            menu_item = ""
            '''menu_item = Drink(
                category_id=category.id,
                price=price,
                volume=weight_or_volume,
                name=name,
                description=description,
                cooking_time_minutes=cooking_time_minutes
            )'''
        elif is_dishes:
            menu_item = Dish(
                category_id=category.id,
                price=price,
                weight=weight_or_volume,
                name=name,
                description=description,
                cooking_time_minutes=cooking_time_minutes
            )
        else:
            raise ValueError("2nd columns name must be 'Вес (гр)' or 'Объем (мл)'")

        menu_items.append(menu_item)

    for category in categories.values():
        category.dishes = [item for item in menu_items if item.category_id == category.id]
        #category.menu_items = [item for item in menu_items if item.category_id == category.id]

    menu.categories = list(categories.values())

    return menu


default_menu = Menu(restaurant_id=3)
menu = parse_menu_from_excel(r"C:\Users\admin\Desktop\fiit\MealMate\frontend\Business\tg_business_bot\exel_tables_work\TableTemplate.dishes_template.xlsx", default_menu)
api_client = ApiClient(Menu)
api_client.post(menu)
