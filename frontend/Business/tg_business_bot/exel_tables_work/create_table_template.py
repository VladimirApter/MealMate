import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from enum import Enum


class TableTemplate(Enum):
    dishes = "dishes"
    drinks = "drinks"


def _create_menu_part_template(template_type, wb):
    grams_str = "(гр)"
    mls_str = "(мл)"

    if template_type is TableTemplate.dishes:
        weight_or_volume_header = f"Вес {grams_str}"
        image_header = "Изображение блюда"
        sheet_title = "Блюда"
    elif template_type is TableTemplate.drinks:
        weight_or_volume_header = f"Объем {mls_str}"
        sheet_title = "Напитки"
        image_header = "Изображение напитка"
    else:
        raise ValueError(f"Unlnown template type: {template_type}. Possible types in TableTemplate")

    common_headers = [
        "№",
        "Категория",
        "Название",
        "Описание",
        "Цена",
        weight_or_volume_header,
        "Ккал",
        f"Белки {grams_str}",
        f"Жиры {grams_str}",
        f"Углеводы {grams_str}",
        "Время приготовления (в минутах)",
        image_header
    ]

    ws = wb.create_sheet(title=sheet_title)

    for col_num, header in enumerate(common_headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    min_width = 8
    extra_width = 10
    for col_num, header in enumerate(common_headers, 1):
        column_letter = get_column_letter(col_num)
        if header == "№":
            ws.column_dimensions[column_letter].width = len(header) + 2
        elif header == "Изображение блюда":
            ws.column_dimensions[column_letter].width = len(header) + 2
        elif header in ["Название", "Категория"]:
            ws.column_dimensions[column_letter].width = len(header) + extra_width
        elif header == "Описание":
            ws.column_dimensions[column_letter].width = len(header) + extra_width * 2
        else:
            width = max(len(header) + 2, min_width)
            ws.column_dimensions[column_letter].width = width

    for row_num in range(2, 101):
        ws.cell(row=row_num, column=1, value=row_num - 1)

    image_row_height = 75
    for row_num in range(2, 101):
        ws.row_dimensions[row_num].height = image_row_height


def create_menu_template():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _create_menu_part_template(TableTemplate.dishes, wb)
    _create_menu_part_template(TableTemplate.drinks, wb)

    current_dir = os.path.dirname(os.path.abspath(__file__))
    result_path = os.path.join(current_dir, "menu_template.xlsx")
    wb.save(result_path)

