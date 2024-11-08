import openpyxl
from time import time


def generate_table_image_tokens(file_path, menu_id):
    image_tokens = []

    wb = openpyxl.load_workbook(file_path)

    for sheet_index, sheet in enumerate(wb.worksheets, 1):
        if sheet.title in ["Блюда", "Напитки"]:
            image_column = None
            for col_num, header in enumerate(sheet[2], 1):
                if header.value in ["Изображение блюда", "Изображение напитка"]:
                    image_column = col_num
                    break

            if image_column:
                for row_num in range(3, sheet.max_row + 1):
                    cell = sheet.cell(row=row_num, column=image_column)
                    if cell.value == "#VALUE!":
                        image_tokens.append(f"{menu_id}_{sheet_index}_{cell.coordinate}_{hash(time())}")

    return image_tokens
