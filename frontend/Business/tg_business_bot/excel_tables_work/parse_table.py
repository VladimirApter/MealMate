import openpyxl

from tg_business_bot.Models.Menu import Menu
from tg_business_bot.Models.Category import Category
from tg_business_bot.Models.Dish import Dish
from tg_business_bot.Models.Drink import Drink
from tg_business_bot.Models.Nutrients import Nutrients
from tg_business_bot.ApiClient import ApiClient
from tg_business_bot.excel_tables_work.extract_images import extract_and_save_images


def parse_menu_from_excel(file_path: str, menu: Menu) -> Menu:
    dish_api_client = ApiClient(Dish)
    drink_api_client = ApiClient(Drink)
    category_api_client = ApiClient(Category)
    menu_api_client = ApiClient(Menu)
    menu.id = menu_api_client.post(menu)

    images_paths = extract_and_save_images(file_path)

    wb = openpyxl.load_workbook(file_path)

    categories = {}
    menu_items = []

    image_index = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        is_drinks = sheet_name == "Напитки"
        is_dishes = sheet_name == "Блюда"

        for row in ws.iter_rows(min_row=3, values_only=True):
            row_has_info = any(row[1:12])
            if not row_has_info:
                continue

            category_name = row[1]
            name = row[2]
            description = row[3]
            price = row[4]
            weight_or_volume = row[5]
            kilocalories = row[6]
            proteins = row[7]
            fats = row[8]
            carbohydrates = row[9]
            cooking_time_minutes = row[10]
            image_value = row[11]
            image_path = None
            if image_value == "#VALUE!":
                image_path = images_paths[image_index]
                image_index += 1

            if category_name not in categories:
                category = Category(menu_id=menu.id, name=category_name)
                category.id = category_api_client.post(category)
                categories[category_name] = category

            category = categories[category_name]

            if is_drinks:
                menu_item = Drink(
                    category_id=category.id,
                    price=price,
                    volume=weight_or_volume,
                    name=name,
                    description=description,
                    cooking_time_minutes=cooking_time_minutes,
                    image_path=image_path
                )
                menu_item_id = drink_api_client.post(menu_item)
            elif is_dishes:
                menu_item = Dish(
                    category_id=category.id,
                    price=price,
                    weight=weight_or_volume,
                    name=name,
                    description=description,
                    cooking_time_minutes=cooking_time_minutes,
                    image_path=image_path
                )
                menu_item_id = dish_api_client.post(menu_item)
            else:
                raise ValueError("Sheet name must be 'Блюда' or 'Напитки'")

            menu_item.nutrients_of_100_grams = Nutrients(menu_item_id, kilocalories, proteins, fats, carbohydrates)

            menu_items.append(menu_item)

    for category in categories.values():
        category.menu_items = [item for item in menu_items if item.category_id == category.id]

    menu.categories = list(categories.values())

    return menu
