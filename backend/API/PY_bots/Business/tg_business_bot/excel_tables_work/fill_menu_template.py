import openpyxl
from tg_business_bot.Models.Menu import Menu
from tg_business_bot.Models.Dish import Dish
#from tg_business_bot.Models.Drink import Drink


def fill_menu_template(template_path, output_file_path, menu: Menu):
    wb = openpyxl.load_workbook(template_path)

    dishes_sheet = wb["Блюда"]
    drinks_sheet = wb["Напитки"]

    row_num = 3
    for category in menu.categories:
        #for item in category.menu_items:
        for item in category.dishes:
            '''if isinstance(item, Drink):
                sheet = drinks_sheet
                weight_or_volume = item.volume'''
            if isinstance(item, Dish):
                sheet = dishes_sheet
                weight_or_volume = item.weight
            else:
                continue

            sheet.cell(row=row_num, column=2, value=category.name)
            sheet.cell(row=row_num, column=3, value=item.name)
            sheet.cell(row=row_num, column=4, value=item.description)
            sheet.cell(row=row_num, column=5, value=item.price)
            sheet.cell(row=row_num, column=6, value=weight_or_volume)
            sheet.cell(row=row_num, column=11, value=item.cooking_time_minutes)

            row_num += 1

    with open(output_file_path, 'w'):
        pass
    wb.save(output_file_path)
