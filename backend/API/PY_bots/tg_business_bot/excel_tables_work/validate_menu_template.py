import openpyxl
from enum import Enum


class CellError(Enum):
    RequiredEmpty = 0
    EmptyTable = 1
    InvalidType = 2


class InvalidCell:
    def __init__(self, cell, error, expected_type=None):
        self.cell = cell
        self.error = error
        self.expected_type = expected_type

    def __str__(self):
        return f"({self.cell.coordinate}, {self.error})"

    def __repr__(self):
        return f"({self.cell.coordinate!r}, {self.error!r})"


def validate_menu_template(file_path):
    wb = openpyxl.load_workbook(file_path)

    invalid_cells_dishes = []
    invalid_cells_drinks = []

    def validate_cell(cell, expected_type, allow_empty=False, invalid_cells_list=None):
        if cell.value is None:
            if not allow_empty:
                invalid_cells_list.append(InvalidCell(cell, CellError.RequiredEmpty))
        else:
            if expected_type == "string":
                if not isinstance(cell.value, str):
                    invalid_cells_list.append(InvalidCell(cell, CellError.InvalidType, expected_type))
            elif expected_type == "number":
                if not isinstance(cell.value, (int, float)):
                    invalid_cells_list.append(InvalidCell(cell, CellError.InvalidType, expected_type))
                elif cell.value < 0:
                    invalid_cells_list.append(InvalidCell(cell, CellError.InvalidType, expected_type))
            elif expected_type == "image":
                if cell.value != "#VALUE!":
                    invalid_cells_list.append(InvalidCell(cell, CellError.InvalidType, expected_type))

    column_types = [
        ("string", False),  # Категория
        ("string", False),  # Название
        ("string", False),  # Описание
        ("number", False),  # Цена
        ("number", False),  # Вес/Объем
        ("number", True),   # Ккал
        ("number", True),   # Белки
        ("number", True),   # Жиры
        ("number", True),   # Углеводы
        ("number", False),  # Время приготовления
        ("image", True)     # Изображение
    ]

    for sheet_name, invalid_cells_list in [("Блюда", invalid_cells_dishes), ("Напитки", invalid_cells_drinks)]:
        ws = wb[sheet_name]
        empty_table_cell = InvalidCell(ws.cell(row=1, column=1), CellError.EmptyTable)

        all_empty = True
        for row_num in range(3, ws.max_row + 1):
            row_empty = True
            for col_num in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    row_empty = False
                    all_empty = False
                    break

            if not row_empty:
                for col_num, (expected_type, allow_empty) in enumerate(column_types, start=2):
                    validate_cell(ws.cell(row=row_num, column=col_num), expected_type, allow_empty, invalid_cells_list)

                proteins_cell = ws.cell(row=row_num, column=8)
                fats_cell = ws.cell(row=row_num, column=9)
                carbs_cell = ws.cell(row=row_num, column=10)

                if (proteins_cell.value is None) != (fats_cell.value is None) or (fats_cell.value is None) != (carbs_cell.value is None):
                    if proteins_cell.value is None:
                        invalid_cells_list.append(InvalidCell(proteins_cell, CellError.RequiredEmpty))
                    if fats_cell.value is None:
                        invalid_cells_list.append(InvalidCell(fats_cell, CellError.RequiredEmpty))
                    if carbs_cell.value is None:
                        invalid_cells_list.append(InvalidCell(carbs_cell, CellError.RequiredEmpty))

        if all_empty:
            invalid_cells_list.append(empty_table_cell)

    if any(cell.error != CellError.EmptyTable for cell in invalid_cells_dishes) \
            or any(cell.error != CellError.EmptyTable for cell in invalid_cells_drinks):
        is_valid = False
    else:
        if any(cell.error == CellError.EmptyTable for cell in invalid_cells_dishes) \
                and any(cell.error == CellError.EmptyTable for cell in invalid_cells_drinks):
            is_valid = False
        else:
            is_valid = True

    return is_valid, invalid_cells_dishes, invalid_cells_drinks


def get_validation_error_messages(invalid_cells_dishes, invalid_cells_drinks):
    def get_error_messages(invalid_cells):
        required_empty_errors = []
        invalid_type_errors = []
        empty_table_error = False

        for invalid_cell in invalid_cells:
            if invalid_cell.error == CellError.RequiredEmpty:
                required_empty_errors.append(invalid_cell.cell.coordinate)
            elif invalid_cell.error == CellError.InvalidType:
                invalid_type_errors.append((invalid_cell.cell.coordinate, invalid_cell.expected_type))
            elif invalid_cell.error == CellError.EmptyTable:
                empty_table_error = True

        return required_empty_errors, invalid_type_errors, empty_table_error

    def get_readable_type(expected_type):
        if expected_type == "string":
            return "строка"
        elif expected_type == "number":
            return "число"
        elif expected_type == "image":
            return "изображение в ячейке"
        return expected_type

    dishes_required_empty, dishes_invalid_type, dishes_empty_table = get_error_messages(invalid_cells_dishes)
    drinks_required_empty, drinks_invalid_type, drinks_empty_table = get_error_messages(invalid_cells_drinks)

    messages = []

    if dishes_empty_table and drinks_empty_table:
        messages.append("Файл пустой.")
    else:
        if dishes_required_empty or dishes_invalid_type:
            messages.append("\nНекорректно заполненные ячейки в таблице Блюда:\n")
        if dishes_required_empty:
            messages.append(f"Пустые ячейки, которые должны быть заполнены:\n{', '.join(dishes_required_empty)}\n")
        if dishes_invalid_type:
            messages.append(f"Неправильные типы данных:\n{', '.join([f'{coordinate} (ожидается {get_readable_type(expected_type)})' for coordinate, expected_type in dishes_invalid_type])}\n")

        if drinks_required_empty or drinks_invalid_type:
            messages.append("\nНекорректно заполненные ячейки в таблице Напитки:\n")
        if drinks_required_empty:
            messages.append(f"Пустые ячейки, которые должны быть заполнены:\n{', '.join(drinks_required_empty)}\n")
        if drinks_invalid_type:
            messages.append(f"Неправильные типы данных:\n{', '.join([f'{coordinate} (ожидается {get_readable_type(expected_type)})' for coordinate, expected_type in drinks_invalid_type])}\n")

    return "\n".join(messages)
