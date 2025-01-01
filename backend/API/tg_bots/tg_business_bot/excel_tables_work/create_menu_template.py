import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from enum import Enum


class TableTemplate(Enum):
    dishes = "dishes"
    drinks = "drinks"


def _create_menu_part_template(template_type, wb):
    grams_str = "(гр)"
    mls_str = "(мл)"

    if template_type is TableTemplate.dishes:
        weight_or_volume_header = f"Вес {grams_str}"
        image_header = "Изображение блюда"
        sheet_title = "Блюда"
    elif template_type is TableTemplate.drinks:
        weight_or_volume_header = f"Объем {mls_str}"
        sheet_title = "Напитки"
        image_header = "Изображение напитка"
    else:
        raise ValueError(f"Unlnown template type: {template_type}. Possible types in TableTemplate")

    common_headers = [
        "№",
        "Категория",
        "Название",
        "Описание",
        "Цена (₽)",
        weight_or_volume_header,
        "Ккал",
        f"Белки {grams_str}",
        f"Жиры {grams_str}",
        f"Углеводы {grams_str}",
        "Время приготовления (мин)",
        image_header
    ]

    ws = wb.create_sheet(title=sheet_title)

    title_cell = ws.cell(row=1, column=1, value=sheet_title)
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    title_cell.border = Border(
        left=Side(style='thin', color='FF000000'),
        right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'),
        bottom=Side(style='thin', color='FF000000')
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(common_headers))

    for col_num, header in enumerate(common_headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True)

    min_width = 8
    extra_width = 10
    for col_num, header in enumerate(common_headers, 1):
        column_letter = get_column_letter(col_num)
        if header == "№":
            ws.column_dimensions[column_letter].width = len(header) + 2
        elif header == image_header:
            ws.column_dimensions[column_letter].width = len(header) + 2
        elif header in ["Название", "Категория"]:
            ws.column_dimensions[column_letter].width = len(header) + extra_width
        elif header == "Описание":
            ws.column_dimensions[column_letter].width = len(header) + extra_width * 2
        else:
            width = max(len(header) + 2, min_width)
            ws.column_dimensions[column_letter].width = width

    for row_num in range(3, 102):
        ws.cell(row=row_num, column=1, value=row_num - 2)

    image_row_height = 75
    for row_num in range(3, 102):
        ws.row_dimensions[row_num].height = image_row_height

    mandatory_columns = [2, 3, 4, 5, 6, 11]  # Категория, Название, Описание, Цена, Вес(Объем), Время приготовления
    optional_columns = [7, 8, 9, 10, 12]  # Ккал, Белки, Жиры, Углеводы, Изображение

    good_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    good_border = Border(left=Side(style='thin', color='FF006100'),
                         right=Side(style='thin', color='FF006100'),
                         top=Side(style='thin', color='FF006100'),
                         bottom=Side(style='thin', color='FF006100'))

    bad_fill = PatternFill(start_color="FFF8CBAD", end_color="FFF8CBAD", fill_type="solid")
    bad_border = Border(left=Side(style='thin', color='FF9C5714'),
                        right=Side(style='thin', color='FF9C5714'),
                        top=Side(style='thin', color='FF9C5714'),
                        bottom=Side(style='thin', color='FF9C5714'))

    for col_num in mandatory_columns:
        column_letter = get_column_letter(col_num)
        for row_num in range(3, 102):
            ws.conditional_formatting.add(f"{column_letter}{row_num}", FormulaRule(formula=[f'AND(ISBLANK({column_letter}{row_num}), COUNTA($B{row_num}:$L{row_num})>0)'], stopIfTrue=True, fill=bad_fill, border=bad_border))
            ws.conditional_formatting.add(f"{column_letter}{row_num}", FormulaRule(formula=[f'NOT(ISBLANK({column_letter}{row_num}))'], stopIfTrue=True, fill=good_fill, border=good_border))

    for col_num in optional_columns:
        column_letter = get_column_letter(col_num)
        for row_num in range(3, 102):
            ws.conditional_formatting.add(f"{column_letter}{row_num}", FormulaRule(formula=[f'NOT(ISBLANK({column_letter}{row_num}))'], stopIfTrue=True, fill=good_fill, border=good_border))

    protein_col = get_column_letter(8)
    fat_col = get_column_letter(9)
    carbs_col = get_column_letter(10)
    for row_num in range(3, 102):
        ws.conditional_formatting.add(f"{protein_col}{row_num}", FormulaRule(formula=[f'AND(ISBLANK({protein_col}{row_num}), OR(NOT(ISBLANK({fat_col}{row_num})), NOT(ISBLANK({carbs_col}{row_num}))))'], stopIfTrue=True, fill=bad_fill, border=bad_border))
        ws.conditional_formatting.add(f"{fat_col}{row_num}", FormulaRule(formula=[f'AND(ISBLANK({fat_col}{row_num}), OR(NOT(ISBLANK({protein_col}{row_num})), NOT(ISBLANK({carbs_col}{row_num}))))'], stopIfTrue=True, fill=bad_fill, border=bad_border))
        ws.conditional_formatting.add(f"{carbs_col}{row_num}", FormulaRule(formula=[f'AND(ISBLANK({carbs_col}{row_num}), OR(NOT(ISBLANK({protein_col}{row_num})), NOT(ISBLANK({fat_col}{row_num}))))'], stopIfTrue=True, fill=bad_fill, border=bad_border))


def create_menu_template():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _create_menu_part_template(TableTemplate.dishes, wb)
    _create_menu_part_template(TableTemplate.drinks, wb)

    current_dir = os.path.dirname(os.path.abspath(__file__))

    tables_path = os.getenv("TABLES_PATH", None)
    result_path = os.path.join(current_dir, "menu_template.xlsx") if tables_path == None else os.path.join(tables_path, "menu_template.xlsx")
    wb.save(result_path)
